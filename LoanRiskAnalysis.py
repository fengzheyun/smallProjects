# -*- coding: utf-8 -*-
#You have been given two datasets. 
#One dataset includes application data for every customer that has been given a loan in a 6 month period.
#The other contains every loan that has been given in this time and whether it has been a good loan or a bad loan. 
#Use the data to identify which new applicants should be given a loan in the future. 
#Discuss any problems or limitations with the data, shortcomings of the analysis and/or additional 
#data elements that might be useful. 
#You should present your findings in a presentation lasting less than 30 minutes including time for questions. 
"""
Written by Zheyun Feng
project for onsite of ZestFinance on 7/17/2015
"""

#%% import libraries
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt

#%% difine data class
class DataItem:
    def __init__(self):
        self.binary = [0]*12;
        self.numeric = [0]*23;
        self.label = 0;

#%% input the data
data_file_name = 'Homework-BusinessAnalysisML.xlsx';
workbook = load_workbook(data_file_name, use_iterators=True)
sheet3 = workbook.get_sheet_names()[2];
worksheet3 = workbook.get_sheet_by_name(sheet3);
sheet2 = workbook.get_sheet_names()[1]
worksheet2 = workbook.get_sheet_by_name(sheet2)

#%% read data label
labels = dict();
for row in list(worksheet3.iter_rows())[1:]:
    col1 =  row[0].internal_value.split('-');
    key = col1[0].lower(); 
    if row[1].internal_value=="Bad":
        labels[key]=labels.get(key,0)-1
    else:
        labels[key]=labels.get(key,0)+1
   
#%% helper functions
def votingWeight( col, labels, workbook ): #zip:15   
    sheet2 = workbook.get_sheet_names()[1]
    worksheet2 = workbook.get_sheet_by_name(sheet2) 
    stat = dict();
    for row in list(worksheet2.iter_rows())[1:]:
        key =  row[0].internal_value;
        if 15==col:
            domain = row[col].internal_value
        elif 16==col:
            domain = row[col].internal_value.split('@')[1];
        if domain not in stat:
            stat[domain] = [0]*2;
        if key not in labels or labels.get( key )==0 :
            continue;   
        if labels.get( key )<0:
            stat[domain][0] += 1;
        else:
            stat[domain][1] += 1;
            
    ratio = dict();
    for key in stat:
        if stat[key][0]+stat[key][1] > 10:
            ratio[key] = 1.0*stat[key][0]/(stat[key][0]+stat[key][1]);
        else:
            ratio[key] = 0.55;
    return ratio;
    
def emailDuration():
    stat = dict();
    stat['3 months or less'] = 1;
    stat['4-12 months'] = 2;
    stat['1 year or more'] = 3;  
    return stat;
    
def bank_residenceDuration():
    stat = dict();
    stat[None] = 0;
    stat['6 months or less'] = 1;
    stat['7-12 months'] = 2;
    stat['1-2 years'] = 3;
    stat['3+ years'] = 4;  
    return stat;
    
def loanPurpose( s ):
    a,b,c,d = False,False,False,False;
    if s=='Rent \/ Mortgage' or s=='School' or s=='Won\'t Say' or s=='Rent' \
        or s=='Pay off loans \/ overdrawn acct' or s=='Medical' or s=='Other':
        a = True;
    if s=='Won\'t Say' or s=='School' or s=='Bills (Home \/ Utilities)' \
        or s=='Bills (Medical)' or s=='Car' or s=='Gifts \/ Leisure' or s=='Other':
        b = True;
    if s=='Won\'t Say' or s=='Rent \/ Mortgage' or s=='Rent' or s=='Bills (Auto)' \
        or s=='Bills (General)' or s=='Car' or s=='Gifts \/ Leisure':
        c = True;
    if s=='Won\'t Say' or s=='Rent \/ Mortgage' or s=='Pay off loans \/ overdrawn acct' \
        or s=='Bills' or s=='Bills (General)' or s=='Bills (Medical)' \
        or s=='Gifts \/ Leisure' or s=='Other':
        d = True;
    return (a,b,c,d);
    
#%% read data features       
ratio_zip = votingWeight( 15, labels, workbook );
ratio_email = votingWeight( 16, labels, workbook );
email_duration = emailDuration();
br_duration = bank_residenceDuration();
 
feature = list();
for row in list(worksheet2.iter_rows())[1:]:
    data = DataItem();
    ib, inu = 0,0;
    key =  row[0].internal_value;
    if  key not in labels or 0==labels.get(key):
        continue;
    data.label = labels.get(key);
    data.numeric[inu] = row[1].internal_value;inu+=1;#amount requested
    data.numeric[inu] = 2011-row[2].internal_value.year;inu+=1;#birth date
    data.binary[ib] = row[4].internal_value;ib+=1;#residence: rent or own
    data.numeric[inu] = row[5].internal_value;inu+=1;#monly rent amount
    data.binary[ib] = row[6].internal_value;ib+=1;#bank account direct deposit    
    data.numeric[inu] = 12*(2011-row[7].internal_value.year)+5-row[7].internal_value.month;inu+=1;
    data.numeric[inu] = row[8].internal_value;inu+=1;#loan duration
    data.numeric[inu] = row[10].internal_value;inu+=1;#number payments
    data.numeric[inu] = row[11].internal_value;inu+=1;#payment amount
    data.numeric[inu] = row[12].internal_value;inu+=1;#amount approved
    data.numeric[inu] = 1.0*data.numeric[7]/data.numeric[0];inu+=1;#approved ratio
    data.numeric[inu] = row[13].internal_value;inu+=1;#duration approved
    if row[14].internal_value == None:#payment amount approved
        data.binary[ib] = False;ib+=1;
        data.numeric[inu] = 0.9*row[11].internal_value; inu+=1;
    else:
        data.binary[ib] = True;ib+=1;
        data.numeric[inu] = row[14].internal_value; inu+=1;
    data.numeric[inu] = ratio_zip.get(row[15].internal_value );inu+=1;#zip code
    data.numeric[inu] = ratio_email.get(row[16].internal_value.split('@')[1] );inu+=1;#email
    data.numeric[inu] = email_duration.get( row[18].internal_value );inu+=1;#email duration
    data.numeric[inu] = br_duration.get( row[19].internal_value );inu+=1;#residence duration
    data.numeric[inu] = br_duration.get( row[20].internal_value );inu+=1;#bank duration
    if row[21].internal_value=='Weekly':#payment frequency
        data.binary[ib] = False;ib+=1;
        data.binary[ib] = False;ib+=1;
    elif row[21].internal_value=='Bi-weekly':
        data.binary[ib] = False;ib+=1;
        data.binary[ib] = True;ib+=1;
    elif row[21].internal_value=='Semi-monthly':
        data.binary[ib] = True;ib+=1;
        data.binary[ib] = False;ib+=1;
    else:
        data.binary[ib] = True;ib+=1;
        data.binary[ib] = True;ib+=1;    
    data.binary[ib] = False;ib+=1;#contaction way
    data.binary[ib] = False;ib+=1;
    data.binary[ib] = False;ib+=1;
    if row[23].internal_value=='Mobile' or row[22].internal_value=='Mobile':
        data.binary[ib-3] = True;
    if row[23].internal_value=='Home' or row[22].internal_value=='Home':
        data.binary[ib-2] = True;
    if row[23].internal_value=='Work' or row[22].internal_value=='Work':
        data.binary[ib-1] = True;
    a,b,c,d = loanPurpose( row[24].internal_value )
    data.binary[ib] = a;ib+=1;
    data.binary[ib] = b;ib+=1;
    data.binary[ib] = c;ib+=1;
    data.binary[ib] = d;ib+=1;
    data.numeric[inu] = 1.0*row[11].internal_value/(row[25].internal_value-row[5].internal_value);inu+=1;
    for i in range(25,31):
        data.numeric[inu] = row[i].internal_value;inu+=1;#duration approved    
    feature.append( data )
    


#%% Present data feature as a matrix, and normalization
rows, cols = len(feature), len(feature[0].binary)+len(feature[0].numeric)
X = np.zeros((rows,cols));
Y = np.zeros((rows,))
i = 0
for dataitem in feature:
    X[i] = dataitem.binary + dataitem.numeric;
    Y[i] = dataitem.label if dataitem.label<=0 else 1 # combine 1 and 2 if label>0
    i += 1
#X_normalized = arrNormalize(X,norm='l2',axis=0)
X_normalized = X - X.min( 0,keepdims=True )   
X_normalized = X_normalized/X_normalized.max(0,keepdims=True)

#%% histogram of bad/good loanees against certain feature
#plt.hist(X[Y==1,17], bins=20, histtype='stepfilled', normed=True, color='b', label='Good')
#plt.hist(X[Y==-1,17], bins=20, histtype='stepfilled', normed=True, color='r', alpha=0.5, label='Bad')
#plt.title("Histogram of good/bad loanees against number of payments")
#plt.xlabel("number of payments")
#plt.ylabel("Number of loanees (%)")
#plt.legend()
#plt.show()

#%% classification
#%% librarues
from sklearn import svm, metrics, linear_model
#from sklearn.linear_model import SGDClassifier
from sklearn.ensemble import AdaBoostClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.lda import LDA
from sklearn.qda import QDA
from sklearn.neighbors import KNeighborsClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.ensemble import ExtraTreesClassifier
from sklearn.ensemble import BaggingClassifier
#from sklearn.preprocessing import normalize as arrNormalize


#%%

NN = 30
Perf_M = []
for k in range(NN):
    Perf_M.append([])
    # split data into training and testing
#%%
    training_perc, testing_perc = [0.8,0.2]
    rng = np.random.permutation(rows)
    idx_train =rng[:int(rows*training_perc)]
    idx_test=rng[int(rows*training_perc):]
    
    training_data,training_label = X_normalized[idx_train], Y[idx_train]
    testing_data,testing_label = X_normalized[idx_test], Y[idx_test]
    
    # %%
    classifiers=[
                svm.SVC(C=100, cache_size=200, class_weight=None, coef0=0.0, degree=3,
                         gamma=0.0, kernel='poly', max_iter=-1, probability=True, random_state=None,
                         shrinking=True, tol=0.001, verbose=False),
                linear_model.LogisticRegression(),
                LDA(),
                QDA(),
                AdaBoostClassifier(DecisionTreeClassifier(max_depth=5), n_estimators=400),
                KNeighborsClassifier(n_neighbors=30),
                BaggingClassifier(DecisionTreeClassifier(max_depth=5), n_estimators=400),
                RandomForestClassifier(max_depth=5, n_estimators=100),
                ExtraTreesClassifier(bootstrap=True, max_depth=5)
                ]
    
    
# training
    idx_clf = 0;
    classifier = classifiers[idx_clf]
    classifier.fit(training_data, training_label)
    
    #testing
    
    predicted_label_test = classifier.predict(testing_data)
    predicted_probability_test=classifier.predict_proba(testing_data)
    ConFusionMat_test=metrics.confusion_matrix(testing_label, predicted_label_test)
    ClfReport_test=metrics.classification_report(testing_label, predicted_label_test)
    accuracy = metrics.accuracy_score(testing_label, predicted_label_test)
#    print "\nClassifier:", classifier
#    print ConFusionMat_test
#    print ClfReport_test
#    print 'accuracy = %f'%accuracy
    
    # %% combination of classifiers
    combine_predicted_probability_test = np.zeros(predicted_probability_test.shape)
    for idx_clf in range(len(classifiers)):
        classifier = classifiers[idx_clf]
        classifier.fit(training_data, training_label)
        
        #testing
        predicted_label_test = classifier.predict(testing_data)
        predicted_probability_test = classifier.predict_proba(testing_data)
        accuracy = metrics.accuracy_score(testing_label, predicted_label_test)
        Perf_M[-1].append(accuracy)
        mean_acc = np.mean([a[idx_clf] for a in Perf_M])
        combine_predicted_probability_test += np.round(predicted_probability_test)*(mean_acc-0.55)
        ConFusionMat_test=metrics.confusion_matrix(testing_label, predicted_label_test)
        ClfReport_test=metrics.classification_report(testing_label, predicted_label_test)
        print '*'*60, "\nClassifier:", classifier
        print ConFusionMat_test
        print ClfReport_test
        print 'accuracy = %f'%accuracy
        
    
    combined_predicted_label_test = classifier.classes_[combine_predicted_probability_test.argmax(axis=1)]
    ConFusionMat_test=metrics.confusion_matrix(testing_label, combined_predicted_label_test)
    ClfReport_test=metrics.classification_report(testing_label, combined_predicted_label_test)
    accuracy = metrics.accuracy_score(testing_label, combined_predicted_label_test)
    print '*-'*30, '\n'*2 + "Combined Classifier:"
    print ConFusionMat_test
    print ClfReport_test
    print 'accuracy = %f'%accuracy
    Perf_M[-1].append(accuracy)

#%%
Perf_M1 = np.array(Perf_M)
for ir in range(Perf_M1.shape[0])[::-1]:
    Perf_M1[ir,:]=np.mean(Perf_M1[:ir+1,:],axis=0)
plt.figure(figsize=(8,5))    
colormap = plt.cm.gist_ncar
plt.gca().set_color_cycle([colormap(ic) for ic in np.linspace(0, 0.9, len(classifiers))])
markers=['*','+','x']*4
for  _s, _x in zip(markers, np.transpose(Perf_M1[:,:-1])):
    plt.plot(_x, marker = _s);

plt.plot(Perf_M1[:,-1],'r-o');
plt.axis([0, Perf_M1.shape[0], 0.5,0.75]);
plt.legend([str(clf).split('(')[0][:16] for clf in classifiers]+['combined'], 
            loc='right',fancybox=True, shadow=True,bbox_to_anchor=[1.35, 0.5]);
plt.title("Performance of classifiers")
plt.xlabel("index of test count")
plt.ylabel("mean accuracy")
plt.savefig('classifier_performance.png')
